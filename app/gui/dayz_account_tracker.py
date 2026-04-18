# app/gui/dayz_account_tracker.py — DayZ Account Tracker
"""DayZ multi-account tracker with CSV/XLSX import/export.

``DayZAccountTracker`` presents a searchable, sortable table of game
accounts with per-row status colouring, multi-select, right-click
context menus, quick-filter chips, bulk operations, and full CRUD via
a modal dialog.  Persistence is delegated to
``app.core.data_persistence.account_manager``.
"""

from __future__ import annotations

import csv
import os
from copy import deepcopy
from datetime import datetime
from typing import Any, Dict, List, Optional, Set

from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QAction, QColor, QFont
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from app.core.data_persistence import account_manager
from app.logs.logger import log_error, log_info, log_warning

__all__ = ["DayZAccountTracker"]

# ── Shared constants ───────────────────────────────────────────────
ACCOUNT_FIELDS = ['account', 'email', 'location', 'value', 'status',
                  'station', 'gear', 'holding', 'loadout', 'needs', 'notes']
TABLE_HEADERS = ['Account', 'Email', 'Location', 'Value', 'Status',
                 'Station', 'Gear', 'Holding', 'Loadout', 'Needs', 'Notes']
STATUS_OPTIONS = ['Ready', 'Blood Infection', 'Storage', 'Dead', 'Offline']
STATION_OPTIONS = [
    'Exploder Kit', 'Pox Kit', 'Raider Kit', 'Geared',
    'PVP/Raid Kit', 'Raider/PvP Kit', 'PvP Kit/Raider Kit',
]
STATUS_COLORS = {
    'Ready':           QColor(76, 175, 80),
    'Blood Infection': QColor(244, 67, 54),
    'Storage':         QColor(255, 152, 0),
    'Dead':            QColor(158, 158, 158),
    'Offline':         QColor(96, 125, 139),
}
STATUS_LABEL_STYLES: Dict[str, str] = {
    k: f"color: #{c.red():02X}{c.green():02X}{c.blue():02X};"
    for k, c in STATUS_COLORS.items()
}

#: Starter template accounts shown when the tracker is empty.
_TEMPLATE_ACCOUNTS: List[Dict[str, str]] = [
    {
        'account': 'Main Character',
        'email': '',
        'location': 'Chernarus — Elektro',
        'value': '',
        'status': 'Ready',
        'station': 'Geared',
        'gear': 'Plate Carrier, Combat Helmet',
        'holding': 'M4-A1, CR-527, Bandages x4',
        'loadout': 'M4 + ACOG / CR-527',
        'needs': 'Ammo, Food',
        'notes': 'Primary account — keep geared at all times',
    },
    {
        'account': 'Alt — Storage Mule',
        'email': '',
        'location': 'Chernarus — Tisy',
        'value': '',
        'status': 'Storage',
        'station': 'Raider Kit',
        'gear': 'Assault Backpack, Plate Carrier',
        'holding': 'NVGs, LAR, KA-M, spare ammo',
        'loadout': 'LAR + KA-M',
        'needs': '-',
        'notes': 'Parked near Tisy mil tents',
    },
    {
        'account': 'Alt — Fresh Spawn',
        'email': '',
        'location': 'Livonia — Coast',
        'value': '',
        'status': 'Dead',
        'station': '',
        'gear': '',
        'holding': '',
        'loadout': '',
        'needs': 'Everything',
        'notes': '',
    },
]

#: Known header synonyms / tag-identifiers for CSV+XLSX column auto-detection.
#:
#: Keys are already-normalized tokens (see :func:`_normalize_header`). The
#: normalizer lowercases, strips surrounding whitespace, collapses runs of
#: punctuation/whitespace to a single space, and trims leading ``#`` / ``@``
#: tag prefixes, so synonyms like ``"E-Mail"``, ``"E Mail"``, ``"#email"``,
#: and ``" Email "`` all collapse onto the same key.
_XLSX_KNOWN_HEADERS: Dict[str, str] = {
    # account / player
    "account": "account", "accounts": "account",
    "name": "account", "player": "account", "username": "account",
    "character": "account", "character name": "account",
    "char": "account", "handle": "account", "gamer tag": "account",
    "user": "account",
    # email
    "email": "email", "e mail": "email", "mail": "email",
    "email address": "email", "contact": "email",
    # location
    "location": "location", "loc": "location", "map": "location",
    "server": "location", "server location": "location",
    "region": "location", "coords": "location", "coordinates": "location",
    "where": "location",
    # status
    "status": "status", "state": "status", "condition": "status",
    # station / kit
    "station": "station", "kit": "station", "role": "station",
    "class": "station", "build": "station",
    # gear
    "gear": "gear", "equipment": "gear", "equip": "gear",
    "armor": "gear", "armour": "gear",
    # holding / inventory
    "holding": "holding", "inventory": "holding", "inv": "holding",
    "items": "holding", "stash": "holding", "carrying": "holding",
    # loadout / weapons
    "loadout": "loadout", "weapons": "loadout", "weapon": "loadout",
    "guns": "loadout", "primary": "loadout",
    # needs
    "needs": "needs", "need": "needs", "wants": "needs",
    "wishlist": "needs", "todo": "needs", "to do": "needs",
    # value / tier
    "value": "value", "tier": "value", "rating": "value",
    "worth": "value", "priority": "value",
    # notes
    "notes": "notes", "note": "notes",
    "comment": "notes", "comments": "notes", "remarks": "notes",
    "description": "notes", "desc": "notes",
}

#: Default column order when no header row is detected — must match the
#: positional export order used by :attr:`ACCOUNT_FIELDS` / :attr:`TABLE_HEADERS`.
_XLSX_DEFAULT_FIELDS: List[str] = list(ACCOUNT_FIELDS)

#: Canonical casing for status values, so case-variant CSV input still
#: renders with the correct color chip in the table.
_STATUS_CANON: Dict[str, str] = {s.lower(): s for s in STATUS_OPTIONS}

#: Canonical casing for station / kit values.
_STATION_CANON: Dict[str, str] = {s.lower(): s for s in STATION_OPTIONS}


def _normalize_header(raw: object) -> str:
    """Return a canonical form of a header cell.

    Lowercases, strips ``#``/``@`` tag prefixes, collapses runs of
    whitespace or punctuation to a single space, and trims surrounding
    whitespace. Non-str input is coerced via :func:`str`.
    """
    import re as _re
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    # strip a single leading tag-identifier prefix ("#email", "@notes", ":loc")
    s = _re.sub(r"^[#@:]+", "", s)
    # strip a UTF-8 BOM if it slipped through (e.g. ``encoding='utf-8'`` read)
    if s.startswith("\ufeff"):
        s = s[1:]
    # collapse any run of whitespace or punctuation to a single space
    s = _re.sub(r"[\s\-_./\\|]+", " ", s).strip()
    return s


def _canon_status(val: str) -> str:
    """Return the canonical-cased status if *val* matches, else *val* as-is."""
    key = (val or "").strip().lower()
    return _STATUS_CANON.get(key, val.strip() if isinstance(val, str) else val)


def _canon_station(val: str) -> str:
    """Return the canonical-cased station if *val* matches, else *val* as-is."""
    key = (val or "").strip().lower()
    return _STATION_CANON.get(key, val.strip() if isinstance(val, str) else val)

# ── QSS constants ──────────────────────────────────────────────────

_TRACKER_QSS: str = """
    QWidget { background-color: #0f1923; color: #e0e0e0; }
    QGroupBox {
        color: #00d9ff; font-weight: bold; font-size: 12px;
        border: 1px solid #1a2a3a; border-radius: 6px;
        margin-top: 10px; padding: 12px 8px 8px 8px;
        background: #0f1923;
    }
    QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; }
    QPushButton {
        background: #16213e; color: #e0e0e0; border: 1px solid #0f3460;
        padding: 6px 12px; border-radius: 4px; font-size: 11px;
    }
    QPushButton:hover { background: #0f3460; color: #00d9ff; }
    QPushButton:pressed { background: #00d9ff; color: #0a0a0a; }
    QLineEdit {
        background: #0a1628; color: #e0e0e0; border: 1px solid #1a2a3a;
        border-radius: 4px; padding: 6px 10px; font-size: 12px;
    }
    QLineEdit:focus { border: 1px solid #00d9ff; }
    QTableWidget {
        background-color: #0f1923; color: #e0e0e0;
        border: 1px solid #1a2a3a; gridline-color: #1a2a3a; font-size: 12px;
    }
    QTableWidget::item:selected { background-color: rgba(0, 217, 255, 0.2); color: #fff; }
    QTableWidget::item:alternate { background-color: #0a1628; }
    QHeaderView::section {
        background-color: #16213e; color: #00d9ff; padding: 6px;
        border: 1px solid #1a2a3a; font-weight: bold; font-size: 11px;
    }
    QLabel { color: #e0e0e0; }
    QComboBox {
        background: #0a1628; color: #e0e0e0; border: 1px solid #1a2a3a;
        border-radius: 4px; padding: 4px 10px;
    }
    QComboBox QAbstractItemView {
        background: #0f1923; color: #e0e0e0;
        selection-background-color: rgba(0, 217, 255, 0.3);
    }
"""

_HEADER_QSS: str = "color: #00d9ff; letter-spacing: 2px; padding: 4px;"

_STAT_QSS: str = "font-weight: bold; font-size: 12px; padding: 4px 8px;"

_CHIP_BASE_QSS: str = (
    "QPushButton { border-radius: 10px; padding: 3px 10px; font-size: 10px;"
    " font-weight: bold; border: 1px solid %s; color: %s; background: transparent; }"
    " QPushButton:hover { background: %s; }"
    " QPushButton:checked { background: %s; color: #fff; border-color: %s; }"
)

_DIALOG_QSS: str = """
    QDialog { background-color: #0f1923; color: #e0e0e0; }
    QLabel { color: #cbd5e1; }
    QLineEdit, QTextEdit, QComboBox {
        background: #0a1628; color: #e0e0e0; border: 1px solid #1a2a3a;
        border-radius: 4px; padding: 6px 10px; font-size: 12px;
    }
    QLineEdit:focus, QTextEdit:focus { border: 1px solid #00d9ff; }
    QComboBox QAbstractItemView {
        background: #0f1923; color: #e0e0e0;
        selection-background-color: rgba(0, 217, 255, 0.3);
    }
    QPushButton {
        background: #16213e; color: #e0e0e0; border: 1px solid #0f3460;
        padding: 8px 18px; border-radius: 4px; font-size: 11px; font-weight: bold;
    }
    QPushButton:hover { background: #0f3460; color: #00d9ff; }
"""


# ── DayZAccountTracker ─────────────────────────────────────────────

class DayZAccountTracker(QWidget):
    """Searchable, sortable account table with CRUD, CSV/XLSX I/O, and bulk ops.

    Falls back to a minimal error label when initialisation fails so the
    rest of the dashboard remains usable.
    """

    def __init__(self, controller: Any = None) -> None:
        super().__init__()
        self.controller: Any = controller
        self.accounts: List[Dict[str, Any]] = []
        self.current_account: Optional[Dict[str, Any]] = None
        self.selected_accounts: Set[str] = set()
        self._active_status_filter: Optional[str] = None

        try:
            self._build_ui()
            self._load_accounts()
        except Exception as e:
            log_error(f"Failed to initialize DayZ Account Tracker: {e}")
            self._create_fallback_ui()

    def _create_fallback_ui(self) -> None:
        """Create a minimal fallback UI if initialisation fails."""
        try:
            layout = QVBoxLayout()
            error_label = QLabel(
                "Account Tracker Error\n\n"
                "Failed to initialize account tracker.\n"
                "Please restart the application."
            )
            error_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(error_label)
            self.setLayout(layout)
        except Exception as e:
            log_error(f"Failed to create fallback UI: {e}")

    # ── UI construction ─────────────────────────────────────────────

    def _build_ui(self) -> None:
        """Assemble the tracker: header + management panel."""
        self.setStyleSheet(_TRACKER_QSS)

        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(12, 12, 12, 12)
        self.setLayout(layout)

        header = QLabel("ACCOUNT TRACKER")
        header.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        header.setStyleSheet(_HEADER_QSS)
        layout.addWidget(header)

        layout.addWidget(self._create_account_panel())

    def _create_account_panel(self) -> QWidget:
        """Build the account management panel (controls + chips + stats + table)."""
        panel = QWidget()
        layout = QVBoxLayout()

        # ── Controls bar ──
        controls_group = QGroupBox("Account Management")
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(4)

        for attr, text, handler in [
            ("add_account_btn", "Add", self._add_account),
            ("edit_account_btn", "Edit", self._edit_account),
            ("duplicate_btn", "Duplicate", self._duplicate_account),
            ("delete_account_btn", "Delete", self._delete_account),
            ("bulk_ops_btn", "Bulk Ops", self._show_bulk_operations),
            ("upload_btn", "Upload", self._upload_accounts),
            ("export_csv_btn", "Export", self._export_csv_accounts),
            ("template_btn", "Template", self._load_template),
            ("clear_table_btn", "Clear All", self._clear_account_table),
            ("refresh_btn", "Refresh", self._refresh_account_table),
        ]:
            btn = QPushButton(text)
            btn.clicked.connect(handler)
            setattr(self, attr, btn)
            controls_layout.addWidget(btn)

        # Search
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search accounts...")
        self.search_input.textChanged.connect(self._on_search_or_filter_changed)
        controls_layout.addWidget(self.search_input)

        controls_group.setLayout(controls_layout)
        layout.addWidget(controls_group)

        # ── Quick-filter status chips ──
        chips_layout = QHBoxLayout()
        chips_layout.setSpacing(6)
        chips_label = QLabel("Filter:")
        chips_label.setStyleSheet("color: #64748b; font-size: 11px; font-weight: bold;")
        chips_layout.addWidget(chips_label)

        self._status_chips: Dict[str, QPushButton] = {}
        for status, color in STATUS_COLORS.items():
            hex_c = f"#{color.red():02X}{color.green():02X}{color.blue():02X}"
            chip = QPushButton(status)
            chip.setCheckable(True)
            chip.setStyleSheet(
                _CHIP_BASE_QSS % (hex_c, hex_c, f"rgba({color.red()},{color.green()},{color.blue()},0.15)",
                                  hex_c, hex_c)
            )
            chip.clicked.connect(lambda checked, s=status: self._on_chip_clicked(s, checked))
            self._status_chips[status] = chip
            chips_layout.addWidget(chip)

        # "All" chip to clear filter
        all_chip = QPushButton("All")
        all_chip.setStyleSheet(
            "QPushButton { border-radius: 10px; padding: 3px 10px; font-size: 10px;"
            " font-weight: bold; border: 1px solid #475569; color: #94a3b8;"
            " background: transparent; }"
            " QPushButton:hover { background: rgba(148,163,184,0.15); }"
        )
        all_chip.clicked.connect(self._clear_status_filter)
        self._all_chip = all_chip
        chips_layout.addWidget(all_chip)

        chips_layout.addStretch()
        layout.addLayout(chips_layout)

        # ── Statistics row ──
        stats_layout = QHBoxLayout()
        stat_defs = [
            ("total_accounts_label", "Total: 0", "#00d9ff"),
            ("ready_accounts_label", "Ready: 0", "#4CAF50"),
            ("blood_infection_label", "Blood Infection: 0", "#F44336"),
            ("storage_label", "Storage: 0", "#FF9800"),
            ("dead_label", "Dead: 0", "#9E9E9E"),
            ("offline_label", "Offline: 0", "#607D8B"),
        ]
        for attr, text, color in stat_defs:
            lbl = QLabel(text)
            lbl.setStyleSheet(f"color: {color}; {_STAT_QSS}")
            setattr(self, attr, lbl)
            stats_layout.addWidget(lbl)
        stats_layout.addStretch()
        layout.addLayout(stats_layout)

        # ── Account table ──
        table_group = QGroupBox("Account Details")
        table_layout = QVBoxLayout()

        self.account_table = QTableWidget()
        self.account_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.account_table.customContextMenuRequested.connect(self._show_context_menu)
        self._setup_account_table()
        table_layout.addWidget(self.account_table)

        table_group.setLayout(table_layout)
        layout.addWidget(table_group)

        # ── Status bar ──
        status_row = QHBoxLayout()
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: #64748b; font-size: 11px;")
        status_row.addWidget(self.status_label)

        self.last_modified_label = QLabel("")
        self.last_modified_label.setStyleSheet("color: #475569; font-size: 10px;")
        self.last_modified_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        status_row.addWidget(self.last_modified_label)
        layout.addLayout(status_row)

        panel.setLayout(layout)
        return panel

    def _setup_account_table(self, accounts_to_show: Optional[List[Dict[str, Any]]] = None) -> None:
        """Populate the account table, optionally with a filtered subset."""
        try:
            accounts = accounts_to_show if accounts_to_show is not None else self.accounts

            # Disconnect signals while rebuilding to prevent stacking
            try:
                self.account_table.itemSelectionChanged.disconnect(self._on_account_selected)
            except (TypeError, RuntimeError):
                pass

            # Disable sorting during population
            self.account_table.setSortingEnabled(False)

            # Add a '#' column at position 0
            display_headers = ['#'] + list(TABLE_HEADERS)
            self.account_table.setColumnCount(len(display_headers))
            self.account_table.setHorizontalHeaderLabels(display_headers)
            self.account_table.setRowCount(len(accounts))
            self.account_table.setAlternatingRowColors(True)
            self.account_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
            self.account_table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
            self.account_table.setWordWrap(True)
            self.account_table.setShowGrid(True)
            self.account_table.setGridStyle(Qt.PenStyle.SolidLine)

            # Status column is offset by 1 because of the '#' column
            status_col = ACCOUNT_FIELDS.index('status') + 1 if 'status' in ACCOUNT_FIELDS else 5

            for row, account in enumerate(accounts):
                # Row number (1-indexed)
                num_item = QTableWidgetItem(str(row + 1))
                num_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                num_item.setFlags(num_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                num_item.setForeground(QColor("#475569"))
                self.account_table.setItem(row, 0, num_item)

                for col, field in enumerate(ACCOUNT_FIELDS):
                    val = account.get(field, '')
                    item = QTableWidgetItem(val)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.account_table.setItem(row, col + 1, item)

                # Color code status
                status_item = self.account_table.item(row, status_col)
                if status_item and status_item.text() in STATUS_COLORS:
                    status_item.setBackground(STATUS_COLORS[status_item.text()])

            # Re-enable sorting and reconnect signals
            self.account_table.setSortingEnabled(True)
            self.account_table.itemSelectionChanged.connect(self._on_account_selected)

            # Auto-resize, but clamp # column narrow
            self.account_table.resizeColumnsToContents()
            self.account_table.setColumnWidth(0, 36)

        except Exception as e:
            log_error(f"Failed to setup account table: {e}")

    # ── Context menu ────────────────────────────────────────────────

    def _show_context_menu(self, position) -> None:
        """Right-click context menu on the account table."""
        menu = QMenu(self)
        menu.setStyleSheet(
            "QMenu { background: #16213e; color: #e0e0e0; border: 1px solid #0f3460; }"
            " QMenu::item:selected { background: #0f3460; color: #00d9ff; }"
        )

        selected_rows = self.account_table.selectionModel().selectedRows()
        if not selected_rows:
            add_action = menu.addAction("Add Account")
            add_action.triggered.connect(self._add_account)
            menu.exec(self.account_table.viewport().mapToGlobal(position))
            return

        # Single-row actions
        if len(selected_rows) == 1:
            edit_action = menu.addAction("Edit Account")
            edit_action.triggered.connect(self._edit_account)
            dup_action = menu.addAction("Duplicate Account")
            dup_action.triggered.connect(self._duplicate_account)
            menu.addSeparator()

        # Status sub-menu (works for single or multi-select)
        status_menu = menu.addMenu("Set Status")
        for status in STATUS_OPTIONS:
            action = status_menu.addAction(status)
            action.triggered.connect(lambda checked, s=status: self._set_selected_status(s))

        menu.addSeparator()

        del_action = menu.addAction(
            f"Delete ({len(selected_rows)} selected)"
            if len(selected_rows) > 1
            else "Delete Account"
        )
        del_action.triggered.connect(self._delete_account)

        menu.exec(self.account_table.viewport().mapToGlobal(position))

    def _set_selected_status(self, new_status: str) -> None:
        """Set status for all selected rows via context menu."""
        try:
            selected_rows = self.account_table.selectionModel().selectedRows()
            if not selected_rows:
                return

            changed = 0
            for idx in selected_rows:
                row = idx.row()
                # Name is in column 1 (after '#')
                name_item = self.account_table.item(row, 1)
                if not name_item:
                    continue
                name = name_item.text()
                for acc in self.accounts:
                    if acc.get('account') == name:
                        acc['status'] = new_status
                        acc['updated_at'] = datetime.now().isoformat()
                        changed += 1
                        break

            if changed:
                self._save_accounts()
                self._refresh_account_table()
                self.status_label.setText(f"Set {changed} account(s) to '{new_status}'")
        except Exception as e:
            log_error(f"Failed to set status via context menu: {e}")

    # ── Quick-filter chips ──────────────────────────────────────────

    def _on_chip_clicked(self, status: str, checked: bool) -> None:
        """Toggle a status filter chip."""
        if checked:
            # Uncheck all other chips
            for s, chip in self._status_chips.items():
                if s != status:
                    chip.setChecked(False)
            self._active_status_filter = status
        else:
            self._active_status_filter = None
        self._on_search_or_filter_changed()

    def _clear_status_filter(self) -> None:
        """Clear all chip filters."""
        self._active_status_filter = None
        for chip in self._status_chips.values():
            chip.setChecked(False)
        self._on_search_or_filter_changed()

    def _on_search_or_filter_changed(self) -> None:
        """Refilter the table when search text or status chip changes."""
        try:
            search_text = self.search_input.text().strip().lower() if hasattr(self, 'search_input') else ""
            filtered = list(self.accounts)

            # Apply status chip filter
            if self._active_status_filter:
                filtered = [a for a in filtered
                            if a.get('status', '') == self._active_status_filter]

            # Apply search text filter
            if search_text:
                filtered = [a for a in filtered
                            if any(search_text in a.get(f, '').lower()
                                   for f in ACCOUNT_FIELDS)]

            self._setup_account_table(filtered)
            self._update_statistics()
        except Exception as e:
            log_error(f"Failed to filter accounts: {e}")

    # ── Account CRUD ─────────────────────────────────────────────────

    def _add_account(self) -> None:
        """Open the account dialog and persist a new record."""
        try:
            account_data = self._show_account_dialog()
            if account_data:
                now = datetime.now()
                account_data['id'] = f"manual_{now.timestamp()}"
                account_data['created_at'] = now.isoformat()
                account_data['updated_at'] = now.isoformat()

                if self._is_duplicate_account(account_data):
                    QMessageBox.warning(self, "Duplicate",
                        f"An account named '{account_data.get('account', '')}' already exists.")
                    return

                account_manager.add_account(account_data)
                self.accounts = account_manager.accounts.copy()
                self._refresh_account_table()
                self.status_label.setText(f"Added: {account_data.get('account', '')}")
                log_info(f"Added new account: {account_data.get('account', '')}")

        except Exception as e:
            log_error(f"Failed to add account: {e}")
            QMessageBox.critical(self, "Error", f"Failed to add account: {e}")

    def _edit_account(self) -> None:
        """Edit the currently selected account via dialog."""
        try:
            if not self.current_account:
                QMessageBox.information(self, "No Selection", "Select an account to edit.")
                return

            account_data = self._show_account_dialog(self.current_account)
            if account_data:
                self.current_account.update(account_data)
                self.current_account['updated_at'] = datetime.now().isoformat()

                account_name = self.current_account.get('account', '')
                account_manager.update_account(account_name, self.current_account)

                self.accounts = account_manager.accounts.copy()
                self._refresh_account_table()
                self.status_label.setText(f"Updated: {self.current_account.get('account', '')}")
                log_info(f"Updated account: {self.current_account.get('account', '')}")

        except Exception as e:
            log_error(f"Failed to edit account: {e}")
            QMessageBox.critical(self, "Error", f"Failed to edit account: {e}")

    def _duplicate_account(self) -> None:
        """Clone the currently selected account with a new name."""
        try:
            if not self.current_account:
                QMessageBox.information(self, "No Selection", "Select an account to duplicate.")
                return

            cloned = deepcopy(self.current_account)
            base_name = cloned.get('account', 'Account')
            cloned['account'] = f"{base_name} (copy)"
            now = datetime.now()
            cloned['id'] = f"clone_{now.timestamp()}"
            cloned['created_at'] = now.isoformat()
            cloned['updated_at'] = now.isoformat()

            # Ensure name uniqueness
            suffix = 1
            while self._is_duplicate_account(cloned):
                suffix += 1
                cloned['account'] = f"{base_name} (copy {suffix})"

            account_manager.add_account(cloned)
            self.accounts = account_manager.accounts.copy()
            self._refresh_account_table()
            self.status_label.setText(f"Duplicated: {cloned['account']}")
            log_info(f"Duplicated account: {base_name} -> {cloned['account']}")

        except Exception as e:
            log_error(f"Failed to duplicate account: {e}")
            QMessageBox.critical(self, "Error", f"Failed to duplicate account: {e}")

    def _delete_account(self) -> None:
        """Delete selected account(s) after confirmation."""
        try:
            selected_rows = self.account_table.selectionModel().selectedRows()
            if not selected_rows:
                QMessageBox.information(self, "No Selection", "Select account(s) to delete.")
                return

            # Gather names
            names = []
            for idx in selected_rows:
                name_item = self.account_table.item(idx.row(), 1)
                if name_item:
                    names.append(name_item.text())

            if not names:
                return

            if len(names) == 1:
                msg = f"Delete account '{names[0]}'?"
            else:
                msg = f"Delete {len(names)} selected accounts?\n\n" + "\n".join(f"  - {n}" for n in names[:10])
                if len(names) > 10:
                    msg += f"\n  ... and {len(names) - 10} more"

            reply = QMessageBox.question(
                self, "Delete Account(s)", msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )

            if reply == QMessageBox.StandardButton.Yes:
                for name in names:
                    account_manager.remove_account(name)

                self.accounts = account_manager.accounts.copy()
                self._refresh_account_table()
                self._clear_account_details()
                self.status_label.setText(f"Deleted {len(names)} account(s)")
                log_info(f"Deleted accounts: {names}")

        except Exception as e:
            log_error(f"Failed to delete account: {e}")
            QMessageBox.critical(self, "Error", f"Failed to delete account: {e}")

    def _show_account_dialog(self, account_data: Optional[Dict[str, Any]] = None) -> Optional[Dict[str, str]]:
        """Present a modal form for creating or editing an account."""
        try:
            is_edit = account_data is not None
            dialog = QDialog(self)
            dialog.setWindowTitle("Edit Account" if is_edit else "New Account")
            dialog.setModal(True)
            dialog.setMinimumWidth(420)
            dialog.setStyleSheet(_DIALOG_QSS)

            layout = QVBoxLayout()
            form = QFormLayout()
            form.setSpacing(8)

            # Text input fields
            _text_fields = [
                ("account_input", "Account Name:", "e.g. Main Character", "account"),
                ("email_input", "Email:", "user@domain.com", "email"),
                ("location_input", "Location:", "e.g. Chernarus — Elektro", "location"),
                ("value_input", "Value:", "e.g. $500, High, etc.", "value"),
            ]
            inputs: Dict[str, QLineEdit] = {}
            for var, label, placeholder, key in _text_fields:
                inp = QLineEdit()
                inp.setPlaceholderText(placeholder)
                if account_data:
                    inp.setText(account_data.get(key, ''))
                form.addRow(label, inp)
                inputs[var] = inp

            # Status dropdown (editable — user can type custom statuses)
            status_combo = QComboBox()
            status_combo.setEditable(True)
            status_combo.addItems(STATUS_OPTIONS)
            if account_data:
                current_status = account_data.get('status', 'Ready')
                index = status_combo.findText(current_status)
                if index >= 0:
                    status_combo.setCurrentIndex(index)
                else:
                    status_combo.setCurrentText(current_status)
            form.addRow("Status:", status_combo)

            # Station dropdown (editable)
            station_combo = QComboBox()
            station_combo.setEditable(True)
            station_combo.addItems(STATION_OPTIONS)
            if account_data:
                current_station = account_data.get('station', '')
                index = station_combo.findText(current_station)
                if index >= 0:
                    station_combo.setCurrentIndex(index)
                else:
                    station_combo.setCurrentText(current_station)
            form.addRow("Station:", station_combo)

            # Remaining text fields
            for var, label, placeholder, key in [
                ("gear_input", "Gear:", "e.g. Plate Carrier, Combat Helmet", "gear"),
                ("loadout_input", "Loadout:", "e.g. M4 + ACOG / CR-527", "loadout"),
                ("needs_input", "Needs:", "e.g. Ammo, Food  (or - for none)", "needs"),
            ]:
                inp = QLineEdit()
                inp.setPlaceholderText(placeholder)
                if account_data:
                    inp.setText(account_data.get(key, ''))
                form.addRow(label, inp)
                inputs[var] = inp

            # Holding (multi-line)
            holding_input = QTextEdit()
            holding_input.setMaximumHeight(60)
            holding_input.setPlaceholderText("Items currently holding")
            if account_data:
                holding_input.setText(account_data.get('holding', ''))
            form.addRow("Holding:", holding_input)

            # Notes (multi-line)
            notes_input = QTextEdit()
            notes_input.setMaximumHeight(60)
            notes_input.setPlaceholderText("Private notes about this account")
            if account_data:
                notes_input.setText(account_data.get('notes', ''))
            form.addRow("Notes:", notes_input)

            layout.addLayout(form)
            layout.addSpacing(8)

            # ── Buttons ──
            button_layout = QHBoxLayout()
            button_layout.addStretch()
            save_btn = QPushButton("Save" if is_edit else "Add Account")
            save_btn.setStyleSheet(
                "QPushButton { background: #0f3460; color: #00d9ff;"
                " border: 1px solid #00d9ff; padding: 8px 24px; }"
                " QPushButton:hover { background: #00d9ff; color: #0a0a0a; }"
            )
            cancel_btn = QPushButton("Cancel")

            save_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)

            button_layout.addWidget(save_btn)
            button_layout.addWidget(cancel_btn)
            layout.addLayout(button_layout)

            dialog.setLayout(layout)

            if dialog.exec() == QDialog.DialogCode.Accepted:
                result = {k.replace('_input', ''): v.text() for k, v in inputs.items()}
                result['status'] = status_combo.currentText()
                result['station'] = station_combo.currentText()
                result['holding'] = holding_input.toPlainText()
                result['notes'] = notes_input.toPlainText()
                return result

            return None

        except Exception as e:
            log_error(f"Account dialog error: {e}")
            return None

    # ── Selection and state ──────────────────────────────────────────

    def _on_account_selected(self) -> None:
        """Handle account selection change in the table."""
        try:
            selected_rows = self.account_table.selectionModel().selectedRows()
            if len(selected_rows) == 1:
                row = selected_rows[0].row()
                name_item = self.account_table.item(row, 1)  # column 1 = Account name
                if name_item:
                    account_name = name_item.text()
                    for account in self.accounts:
                        if account.get('account') == account_name:
                            self.current_account = account
                            updated = account.get('updated_at', '')
                            if updated:
                                try:
                                    dt = datetime.fromisoformat(updated)
                                    self.last_modified_label.setText(
                                        f"Last modified: {dt.strftime('%b %d, %Y %I:%M %p')}"
                                    )
                                except ValueError:
                                    self.last_modified_label.setText("")
                            self.status_label.setText(f"Selected: {account_name}")
                            return
                    log_warning(f"Account {account_name} not found in accounts data")
            elif len(selected_rows) > 1:
                self.current_account = None
                self.status_label.setText(f"{len(selected_rows)} accounts selected")
                self.last_modified_label.setText("")
            else:
                self.current_account = None
                self.status_label.setText("Ready")
                self.last_modified_label.setText("")
        except Exception as e:
            log_error(f"Failed to handle account selection: {e}")

    def _clear_account_details(self) -> None:
        """Reset the current selection state."""
        self.current_account = None
        self.status_label.setText("Ready")
        self.last_modified_label.setText("")

    # ── Persistence ─────────────────────────────────────────────────

    def _load_accounts(self) -> None:
        """Load accounts from the persistence layer.

        When the tracker is empty (first launch), auto-loads the
        starter template so the user has example rows to work with.
        """
        try:
            self.accounts = account_manager.accounts.copy()
            # Backfill 'notes' field for accounts that predate this field
            for acc in self.accounts:
                if 'notes' not in acc:
                    acc['notes'] = ''
            if not self.accounts:
                self._apply_template(silent=True)
            self._refresh_account_table()
            self._update_statistics()
            log_info(f"Loaded {len(self.accounts)} accounts")
        except Exception as e:
            log_error(f"Failed to load accounts: {e}")
            self.accounts = []

    def _save_accounts(self) -> None:
        """Flush the local account list back to the persistence layer."""
        try:
            account_manager.accounts[:] = self.accounts.copy()
            account_manager.save_changes(account_manager.accounts)
        except Exception as e:
            log_error(f"Failed to save accounts: {e}")

    def _load_template(self) -> None:
        """Load the starter template, replacing existing accounts after confirmation."""
        try:
            if self.accounts:
                reply = QMessageBox.question(
                    self, "Load Template",
                    "This will replace all existing accounts with the starter template.\nContinue?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if reply != QMessageBox.StandardButton.Yes:
                    return
                account_manager.accounts.clear()

            self._apply_template(silent=False)
            self._refresh_account_table()
            self._update_statistics()
        except Exception as e:
            log_error(f"Failed to load template: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load template: {e}")

    def _apply_template(self, *, silent: bool = False) -> None:
        """Insert template rows into the account manager."""
        now = datetime.now().isoformat()
        for idx, tmpl in enumerate(_TEMPLATE_ACCOUNTS):
            data = dict(tmpl)
            data.update(id=f"template_{idx + 1}", created_at=now, updated_at=now)
            account_manager.accounts.append(data)
        account_manager.save_changes(account_manager.accounts)
        self.accounts = account_manager.accounts.copy()
        if not silent:
            self.status_label.setText(f"Loaded template with {len(_TEMPLATE_ACCOUNTS)} example accounts")
            QMessageBox.information(self, "Template Loaded",
                f"Loaded {len(_TEMPLATE_ACCOUNTS)} example accounts.\n\n"
                "Edit or replace them with your own accounts.")
        log_info(f"Template loaded: {len(_TEMPLATE_ACCOUNTS)} accounts")

    def _refresh_after_import(self) -> None:
        """Common post-import refresh: save, rebuild table, update stats."""
        try:
            account_manager.save_changes(account_manager.accounts)
        except Exception as e:
            log_error(f"Failed to save imports: {e}")
        self.account_table.clearContents()
        self.account_table.setRowCount(0)
        self.accounts = account_manager.accounts.copy()
        self._setup_account_table(self.accounts)
        self._update_statistics()
        self.account_table.resizeColumnsToContents()
        self.account_table.setColumnWidth(0, 36)

    # ── Table refresh / filter ────────────────────────────────────────

    def _refresh_account_table(self) -> None:
        """Rebuild the table from the current account list."""
        try:
            self._on_search_or_filter_changed()
        except Exception as e:
            log_error(f"Failed to refresh account table: {e}")

    def _clear_account_table(self) -> None:
        """Clear all accounts after user confirmation."""
        try:
            reply = QMessageBox.question(
                self, "Clear All Accounts",
                "Are you sure you want to clear all accounts?\nThis action cannot be undone.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )

            if reply == QMessageBox.StandardButton.Yes:
                account_manager.accounts.clear()
                account_manager.save_changes(account_manager.accounts)
                self.accounts = account_manager.accounts.copy()
                self._refresh_account_table()
                self.status_label.setText("All accounts cleared")
                log_info("All accounts cleared")
        except Exception as e:
            log_error(f"Failed to clear accounts: {e}")
            QMessageBox.critical(self, "Error", f"Failed to clear accounts: {e}")

    def _filter_accounts(self, search_text: str) -> None:
        """Filter the table rows by a case-insensitive search term."""
        self._on_search_or_filter_changed()

    # ── Statistics ──────────────────────────────────────────────────

    def _update_statistics(self) -> None:
        """Recompute and display account status counts."""
        try:
            from collections import Counter
            counts = Counter(a.get('status', '') for a in self.accounts)
            self.total_accounts_label.setText(f"Total: {len(self.accounts)}")
            for attr, status in [("ready_accounts_label", "Ready"),
                                 ("blood_infection_label", "Blood Infection"),
                                 ("storage_label", "Storage"), ("dead_label", "Dead"),
                                 ("offline_label", "Offline")]:
                getattr(self, attr).setText(f"{status}: {counts.get(status, 0)}")
        except Exception as e:
            log_error(f"Failed to update statistics: {e}")

    # ── Import / Export ─────────────────────────────────────────────

    def _upload_accounts(self) -> None:
        """Prompt for a CSV or XLSX file and import its rows."""
        try:
            from PyQt6.QtWidgets import QFileDialog

            file_path, _ = QFileDialog.getOpenFileName(
                self, "Select Account File", "",
                "All Supported (*.xlsx *.xls *.csv);;Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)",
            )

            if file_path:
                ext = os.path.splitext(file_path)[1].lower()
                if ext in ('.xlsx', '.xls'):
                    self._process_xlsx_file(file_path)
                elif ext == '.csv':
                    self._process_csv_file(file_path)
                else:
                    QMessageBox.warning(self, "Unsupported File",
                        f"File type '{ext}' is not supported. Use .xlsx or .csv files.")

        except Exception as e:
            log_error(f"Failed to upload accounts: {e}")
            QMessageBox.critical(self, "Error", f"Failed to upload accounts: {e}")

    def _process_xlsx_file(self, file_path: str) -> None:
        """Import accounts from an XLSX workbook with auto-detected headers."""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            # --- Detect header row and column mapping ---
            header_row_idx: Optional[int] = None
            header_map: Dict[int, str] = {}

            for row_idx in range(1, min(6, sheet.max_row + 1)):
                row_vals = []
                for col_idx in range(1, sheet.max_column + 1):
                    cell_val = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_val is not None:
                        row_vals.append((col_idx, _normalize_header(cell_val)))

                matches = {}
                for col_idx, val in row_vals:
                    if val and val in _XLSX_KNOWN_HEADERS:
                        matches[col_idx] = _XLSX_KNOWN_HEADERS[val]

                if len(matches) >= 2:
                    header_row_idx = row_idx
                    header_map = matches
                    break

            if header_row_idx is None:
                # No recognizable header row — positional fallback.
                # We start at row 1 so that row 1 is treated as data, not header.
                header_row_idx = 0
                for col_idx in range(1, min(len(_XLSX_DEFAULT_FIELDS) + 1, sheet.max_column + 1)):
                    header_map[col_idx] = _XLSX_DEFAULT_FIELDS[col_idx - 1]

            log_info(f"XLSX header detected at row {header_row_idx}: {header_map}")

            # --- Parse account rows ---
            accounts_imported = 0
            accounts_skipped = 0
            bases_data = []

            # Detect where "Bases" section starts
            bases_start_row = None
            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                cell_a = sheet.cell(row=row_idx, column=1).value
                if cell_a and 'base' in str(cell_a).strip().lower():
                    cell_b = sheet.cell(row=row_idx, column=2).value
                    if cell_b is None or _normalize_header(cell_b) in _XLSX_KNOWN_HEADERS:
                        bases_start_row = row_idx
                        log_info(f"XLSX bases section detected at row {bases_start_row}")
                        break

            account_end_row = bases_start_row if bases_start_row else sheet.max_row + 1

            # When header_row_idx == 0 (positional fallback), start reading at row 1.
            data_start_row = max(1, header_row_idx + 1)

            for row_idx in range(data_start_row, account_end_row):
                try:
                    row_data = {}
                    is_empty = True
                    for col_idx, field in header_map.items():
                        cell_val = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_val is not None:
                            val = str(cell_val).strip()
                            if val:
                                is_empty = False
                            row_data[field] = val
                        else:
                            row_data[field] = ''

                    if is_empty:
                        continue

                    account_name = row_data.get('account', '').strip()
                    if not account_name or len(account_name) < 2:
                        accounts_skipped += 1
                        continue

                    # Canonicalize status / station casing so color chips still match.
                    if row_data.get('status'):
                        row_data['status'] = _canon_status(row_data['status'])
                    if row_data.get('station'):
                        row_data['station'] = _canon_station(row_data['station'])

                    now = datetime.now().isoformat()
                    account_data = {f: row_data.get(f, 'Ready' if f == 'status' else '')
                                    for f in ACCOUNT_FIELDS}
                    account_data.update(last_seen=now, created_date=now,
                                        id=f"xlsx_{accounts_imported + 1}_{datetime.now().timestamp()}")

                    if self._try_import_account(account_data):
                        accounts_imported += 1
                    else:
                        accounts_skipped += 1

                except Exception as e:
                    accounts_skipped += 1
                    log_error(f"XLSX error at row {row_idx}: {e}")
                    continue

            # --- Parse bases section if present ---
            bases_imported = 0
            if bases_start_row:
                bases_header_map = {}
                row_vals = []
                for col_idx in range(1, sheet.max_column + 1):
                    cell_val = sheet.cell(row=bases_start_row, column=col_idx).value
                    if cell_val is not None:
                        row_vals.append((col_idx, str(cell_val).strip().lower()))

                base_headers = {
                    'base': 'base_name', 'bases': 'base_name', 'name': 'base_name',
                    'code': 'base_code',
                    'location': 'base_location', 'loc': 'base_location',
                    'status': 'base_status', 'notes': 'base_notes',
                }
                for col_idx, val in row_vals:
                    if val in base_headers:
                        bases_header_map[col_idx] = base_headers[val]

                for row_idx in range(bases_start_row + 1, sheet.max_row + 1):
                    base_row = {}
                    is_empty = True
                    for col_idx in range(1, sheet.max_column + 1):
                        cell_val = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_val is not None and str(cell_val).strip():
                            is_empty = False
                            field = bases_header_map.get(col_idx, f'col_{col_idx}')
                            base_row[field] = str(cell_val).strip()
                    if not is_empty:
                        bases_data.append(base_row)
                        bases_imported += 1

                log_info(f"XLSX bases imported: {bases_imported}")

            self._refresh_after_import()

            msg = (f"Import completed!\n\n"
                   f"Accounts imported: {accounts_imported}\n"
                   f"Accounts skipped: {accounts_skipped}\n"
                   f"Total accounts: {len(self.accounts)}")
            if bases_imported > 0:
                msg += f"\n\nBases found: {bases_imported}"
            QMessageBox.information(self, "XLSX Import Complete", msg)
            log_info(f"XLSX import done: {accounts_imported} imported, {accounts_skipped} skipped, {bases_imported} bases")

        except ImportError:
            log_error("openpyxl not installed")
            QMessageBox.critical(self, "Missing Library",
                "openpyxl is required for XLSX import.\n\nInstall it with:\npip install openpyxl")
        except Exception as e:
            log_error(f"Failed to process XLSX file: {e}")
            QMessageBox.critical(self, "Error", f"Failed to process XLSX file:\n{e}")

    def _process_csv_file(self, file_path: str) -> None:
        """Import accounts from a CSV file with header-based column mapping.

        Handles:
            * UTF-8 BOM (``encoding='utf-8-sig'``) so Excel-exported CSVs
              round-trip cleanly.
            * Delimiter sniffing (comma, semicolon, tab, pipe) via
              :class:`csv.Sniffer` with a comma fallback.
            * Expanded header synonyms / tag-identifiers via
              :func:`_normalize_header` + :data:`_XLSX_KNOWN_HEADERS`.
            * Positional fallback when no header row is recognized,
              using :data:`_XLSX_DEFAULT_FIELDS` (which now matches
              :data:`ACCOUNT_FIELDS` exactly).
            * Canonical casing for ``status`` / ``station`` values so
              ``"ready"`` from a foreign CSV still renders with the
              correct color chip.
        """
        try:
            accounts_imported = 0
            accounts_skipped = 0
            unmapped_headers: List[str] = []

            # --- Delimiter sniff on a small probe buffer ---
            with open(file_path, 'r', newline='', encoding='utf-8-sig',
                      errors='replace') as probe:
                sample = probe.read(4096)
            dialect_delim = ','
            if sample.strip():
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                    dialect_delim = dialect.delimiter
                except csv.Error:
                    dialect_delim = ','

            with open(file_path, 'r', newline='', encoding='utf-8-sig',
                      errors='replace') as file:
                reader = csv.DictReader(file, delimiter=dialect_delim)

                csv_col_map: Dict[str, str] = {}
                if reader.fieldnames:
                    for col_name in reader.fieldnames:
                        if col_name is None:
                            continue
                        norm = _normalize_header(col_name)
                        if norm in _XLSX_KNOWN_HEADERS:
                            csv_col_map[col_name] = _XLSX_KNOWN_HEADERS[norm]
                        else:
                            unmapped_headers.append(str(col_name))

                # Positional fallback: if we recognized no headers but the
                # file has a row, assume the canonical ACCOUNT_FIELDS order.
                positional_fallback = False
                if not csv_col_map and reader.fieldnames:
                    log_warning(
                        f"CSV headers unrecognized; falling back to positional "
                        f"mapping. Unmapped: {reader.fieldnames}"
                    )
                    positional_fallback = True
                    # The DictReader already consumed the first row as headers;
                    # treat those header cells as the first data row.
                    first_row_vals = list(reader.fieldnames or [])
                else:
                    first_row_vals = None

                def _row_iter():
                    if positional_fallback and first_row_vals:
                        yield 1, first_row_vals  # synthetic row from headers
                    for idx, r in enumerate(reader, start=2):
                        yield idx, r

                for row_num, row in _row_iter():
                    try:
                        account_data: Dict[str, str] = {f: '' for f in ACCOUNT_FIELDS}
                        account_data['status'] = 'Ready'

                        if positional_fallback and isinstance(row, list):
                            for col_idx, val in enumerate(row):
                                if col_idx >= len(_XLSX_DEFAULT_FIELDS):
                                    break
                                field = _XLSX_DEFAULT_FIELDS[col_idx]
                                if val is not None and str(val).strip():
                                    account_data[field] = str(val).strip()
                        else:
                            for csv_col, field in csv_col_map.items():
                                val = row.get(csv_col, '') if isinstance(row, dict) else ''
                                if val:
                                    account_data[field] = str(val).strip()

                        # Canonicalize status / station casing.
                        if account_data.get('status'):
                            account_data['status'] = _canon_status(account_data['status'])
                        if account_data.get('station'):
                            account_data['station'] = _canon_station(account_data['station'])

                        now = datetime.now().isoformat()
                        account_data.update(
                            last_seen=now, created_date=now,
                            id=f"imported_{accounts_imported + 1}_{datetime.now().timestamp()}",
                        )

                        if self._try_import_account(account_data):
                            accounts_imported += 1
                        else:
                            accounts_skipped += 1

                    except Exception as e:
                        accounts_skipped += 1
                        log_error(f"Error processing row {row_num}: {e}")
                        continue

            self._refresh_after_import()

            detail = (f"Delimiter: {dialect_delim!r}\n"
                      f"Accounts imported: {accounts_imported}\n"
                      f"Accounts skipped: {accounts_skipped}\n\n"
                      f"Total accounts: {len(self.accounts)}")
            if unmapped_headers:
                detail += f"\n\nUnmapped headers (ignored): {', '.join(unmapped_headers)}"
            QMessageBox.information(self, "CSV Import Complete", detail)
            log_info(
                f"CSV import completed: {accounts_imported} imported, "
                f"{accounts_skipped} skipped, delim={dialect_delim!r}, "
                f"unmapped={unmapped_headers}"
            )

        except Exception as e:
            log_error(f"Failed to process CSV file: {e}")
            QMessageBox.critical(self, "Error", f"Failed to process CSV file:\n{e}")

    def _try_import_account(self, account_data: dict) -> bool:
        """Validate, dedup, and add a single account. Returns True if imported.

        Only adds to ``account_manager`` (the source of truth).
        ``self.accounts`` is refreshed from the manager after the
        import batch completes via ``_refresh_after_import``.
        """
        if not self._validate_account_data(account_data):
            return False
        if self._is_duplicate_account(account_data):
            log_warning(f"Skipped duplicate: {account_data.get('account', '')}")
            return False
        try:
            account_manager.accounts.append(account_data)
        except Exception as e:
            log_warning(f"Failed to add account to manager: {e}")
            return False
        log_info(f"Imported account: {account_data['account']}")
        return True

    # ── Validation helpers ──────────────────────────────────────────

    def _validate_account_data(self, data: Dict[str, str]) -> bool:
        """Return *True* if *data* contains a valid account name."""
        try:
            account = data.get('account', '').strip()
            if not account or len(account) < 2:
                log_warning(f"Validation failed: account too short: '{account}'")
                return False
            if account.isspace() or not any(c.isalnum() for c in account):
                log_warning(f"Validation failed: no alphanumeric chars: '{account}'")
                return False
            email = data.get('email', '').strip()
            if email and '@' not in email:
                log_warning(f"Validation failed: bad email: '{email}'")
                return False
            return True
        except Exception as e:
            log_error(f"Error in account validation: {e}")
            return False

    def _is_duplicate_account(self, account_data: Dict[str, Any]) -> bool:
        """Return *True* if an account with the same name already exists."""
        try:
            new_name = account_data.get('account', '').strip().lower()
            if not new_name:
                return False
            for account in account_manager.accounts:
                if account.get('account', '').strip().lower() == new_name:
                    return True
            return False
        except Exception as e:
            log_error(f"Failed to check for duplicate account: {e}")
            return False

    def _export_csv_accounts(self) -> None:
        """Export accounts to a CSV or XLSX file chosen via a save dialog."""
        try:
            from PyQt6.QtWidgets import QFileDialog

            if not self.accounts:
                QMessageBox.information(self, "No Accounts", "No accounts to export.")
                return

            file_path, selected_filter = QFileDialog.getSaveFileName(
                self, "Export Accounts",
                f"dayz_accounts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)",
            )

            if file_path:
                ext = os.path.splitext(file_path)[1].lower()

                if ext == '.xlsx':
                    self._export_xlsx(file_path)
                else:
                    with open(file_path, 'w', newline='', encoding='utf-8') as file:
                        writer = csv.writer(file)
                        writer.writerow(TABLE_HEADERS)
                        for account in self.accounts:
                            writer.writerow([account.get(f, '') for f in ACCOUNT_FIELDS])

                fmt = "XLSX" if ext == '.xlsx' else "CSV"
                self.status_label.setText(f"Exported {len(self.accounts)} accounts to {fmt}")
                QMessageBox.information(self, "Export Complete",
                    f"Exported {len(self.accounts)} accounts to {fmt}.")
                log_info(f"Exported {len(self.accounts)} accounts to {fmt}")

        except Exception as e:
            log_error(f"Failed to export accounts: {e}")
            QMessageBox.critical(self, "Error", f"Failed to export accounts: {e}")

    def _export_xlsx(self, file_path: str) -> None:
        """Write accounts to a styled XLSX workbook with frozen header row."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Accounts"

            headers = list(TABLE_HEADERS)
            fields = list(ACCOUNT_FIELDS)

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill('solid', fgColor="16213E")
            header_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin', color='1A2A3A'),
                right=Side(style='thin', color='1A2A3A'),
                top=Side(style='thin', color='1A2A3A'),
                bottom=Side(style='thin', color='1A2A3A'),
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            status_fills = {k: PatternFill('solid', fgColor=f'{c.red():02X}{c.green():02X}{c.blue():02X}')
                           for k, c in STATUS_COLORS.items()}
            status_col_idx = fields.index('status') + 1 if 'status' in fields else None

            for row_idx, account in enumerate(self.accounts, 2):
                for col_idx, field in enumerate(fields, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=account.get(field, ''))
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=(field in ('holding', 'notes')))

                if status_col_idx:
                    status_val = account.get('status', '')
                    status_cell = ws.cell(row=row_idx, column=status_col_idx)
                    if status_val in status_fills:
                        status_cell.fill = status_fills[status_val]
                        status_cell.font = Font(bold=True, color="FFFFFF")

            for col_idx in range(1, len(headers) + 1):
                max_len = len(headers[col_idx - 1])
                for row_idx in range(2, len(self.accounts) + 2):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

            ws.freeze_panes = 'A2'
            wb.save(file_path)
            log_info(f"XLSX export saved to {file_path}")

        except ImportError:
            QMessageBox.critical(self, "Missing Library",
                "openpyxl is required for XLSX export.\n\nInstall it with:\npip install openpyxl")
        except Exception as e:
            log_error(f"XLSX export failed: {e}")
            raise

    # ── Bulk operations ─────────────────────────────────────────────

    def _show_bulk_operations(self) -> None:
        """Present a dialog for batch status changes, deletion, or filtered ops."""
        try:
            if not self.accounts:
                QMessageBox.information(self, "No Accounts", "No accounts available for bulk operations.")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle("Bulk Operations")
            dialog.setModal(True)
            dialog.setMinimumWidth(380)
            dialog.setStyleSheet(_DIALOG_QSS)

            layout = QVBoxLayout()
            layout.setSpacing(10)

            # ── Scope selector ──
            scope_label = QLabel("Apply to:")
            scope_combo = QComboBox()
            scope_combo.addItems(["All Accounts", "Selected Accounts", "Filtered by Status"])
            layout.addWidget(scope_label)
            layout.addWidget(scope_combo)

            # Status filter (for "Filtered by Status" scope)
            filter_label = QLabel("Where status is:")
            filter_combo = QComboBox()
            filter_combo.addItems(STATUS_OPTIONS)
            filter_label.hide()
            filter_combo.hide()
            layout.addWidget(filter_label)
            layout.addWidget(filter_combo)

            # ── Operation selector ──
            op_label = QLabel("Operation:")
            op_combo = QComboBox()
            op_combo.addItems([
                "Change Status",
                "Set Location",
                "Clear Notes",
                "Delete Accounts",
                "Export Matching",
            ])
            layout.addWidget(op_label)
            layout.addWidget(op_combo)

            # ── Parameter input ──
            param_label = QLabel("New Value:")
            param_combo = QComboBox()
            param_combo.setEditable(True)
            param_combo.addItems(STATUS_OPTIONS)
            layout.addWidget(param_label)
            layout.addWidget(param_combo)

            param_line = QLineEdit()
            param_line.setPlaceholderText("Enter value...")
            param_line.hide()
            layout.addWidget(param_line)

            # ── Dynamic visibility ──
            def _update_visibility():
                op = op_combo.currentText()
                scope = scope_combo.currentText()

                filter_label.setVisible(scope == "Filtered by Status")
                filter_combo.setVisible(scope == "Filtered by Status")

                needs_status = (op == "Change Status")
                needs_text = (op == "Set Location")
                needs_nothing = op in ("Clear Notes", "Delete Accounts", "Export Matching")

                param_label.setVisible(not needs_nothing)
                param_combo.setVisible(needs_status)
                param_line.setVisible(needs_text)

            scope_combo.currentTextChanged.connect(lambda _: _update_visibility())
            op_combo.currentTextChanged.connect(lambda _: _update_visibility())
            _update_visibility()

            # ── Buttons ──
            btn_layout = QHBoxLayout()
            btn_layout.addStretch()
            apply_btn = QPushButton("Apply")
            apply_btn.setStyleSheet(
                "QPushButton { background: #0f3460; color: #00d9ff;"
                " border: 1px solid #00d9ff; padding: 8px 24px; }"
                " QPushButton:hover { background: #00d9ff; color: #0a0a0a; }"
            )
            cancel_btn = QPushButton("Cancel")
            apply_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)
            btn_layout.addWidget(apply_btn)
            btn_layout.addWidget(cancel_btn)
            layout.addLayout(btn_layout)

            dialog.setLayout(layout)

            if dialog.exec() != QDialog.DialogCode.Accepted:
                return

            # ── Resolve target accounts ──
            scope = scope_combo.currentText()
            if scope == "Selected Accounts":
                selected_rows = self.account_table.selectionModel().selectedRows()
                names = set()
                for idx in selected_rows:
                    item = self.account_table.item(idx.row(), 1)
                    if item:
                        names.add(item.text())
                targets = [a for a in self.accounts if a.get('account') in names]
            elif scope == "Filtered by Status":
                filter_status = filter_combo.currentText()
                targets = [a for a in self.accounts if a.get('status') == filter_status]
            else:
                targets = list(self.accounts)

            if not targets:
                QMessageBox.information(self, "No Matches", "No accounts match the selected scope.")
                return

            operation = op_combo.currentText()
            now = datetime.now().isoformat()

            if operation == "Change Status":
                new_status = param_combo.currentText()
                for acc in targets:
                    acc['status'] = new_status
                    acc['updated_at'] = now
                self._save_accounts()
                self._refresh_account_table()
                self.status_label.setText(f"Set {len(targets)} account(s) to '{new_status}'")

            elif operation == "Set Location":
                new_loc = param_line.text().strip()
                if new_loc:
                    for acc in targets:
                        acc['location'] = new_loc
                        acc['updated_at'] = now
                    self._save_accounts()
                    self._refresh_account_table()
                    self.status_label.setText(f"Set location for {len(targets)} account(s)")

            elif operation == "Clear Notes":
                for acc in targets:
                    acc['notes'] = ''
                    acc['updated_at'] = now
                self._save_accounts()
                self._refresh_account_table()
                self.status_label.setText(f"Cleared notes for {len(targets)} account(s)")

            elif operation == "Delete Accounts":
                reply = QMessageBox.question(
                    self, "Confirm Bulk Delete",
                    f"Delete {len(targets)} account(s)? This cannot be undone.",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if reply == QMessageBox.StandardButton.Yes:
                    names_to_remove = {a.get('account') for a in targets}
                    for name in names_to_remove:
                        account_manager.remove_account(name)
                    self.accounts = account_manager.accounts.copy()
                    self._refresh_account_table()
                    self.status_label.setText(f"Deleted {len(names_to_remove)} account(s)")

            elif operation == "Export Matching":
                self._export_subset(targets)

        except Exception as e:
            log_error(f"Failed to show bulk operations: {e}")
            QMessageBox.critical(self, "Error", f"Failed to show bulk operations: {e}")

    def _export_subset(self, subset: List[Dict[str, Any]]) -> None:
        """Export a subset of accounts to XLSX via save dialog."""
        try:
            from PyQt6.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export Accounts",
                f"dayz_accounts_filtered_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)",
            )
            if not file_path:
                return

            ext = os.path.splitext(file_path)[1].lower()
            if ext == '.xlsx':
                # Temporarily swap accounts for export
                orig = self.accounts
                self.accounts = subset
                self._export_xlsx(file_path)
                self.accounts = orig
            else:
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(TABLE_HEADERS)
                    for acc in subset:
                        writer.writerow([acc.get(field, '') for field in ACCOUNT_FIELDS])

            fmt = "XLSX" if ext == '.xlsx' else "CSV"
            QMessageBox.information(self, "Export Complete",
                f"Exported {len(subset)} account(s) to {fmt}.")
        except Exception as e:
            log_error(f"Failed to export subset: {e}")
            QMessageBox.critical(self, "Error", f"Export failed: {e}")
